from flask import Flask, request, jsonify
from flask_cors import CORS
import os
from openpyxl import load_workbook

app = Flask(__name__)
# 퍼블리시된 Webflow 도메인 허용
CORS(app, origins=["https://fast-track-estate.webflow.io"])

# Excel 파일 경로 지정
EXCEL_FILE_PATH = os.getenv("EXCEL_PATH", "기본정보/Basic_index.xlsx")
SHEET_NAME = "head_basic"

@app.route("/search", methods=["GET"])
def search():
    address = request.args.get("address")
    if not address:
        return jsonify({"error": "주소가 없습니다"})

    try:
        # 서버 기준 상대 경로 처리
        excel_path = os.path.join(os.path.dirname(__file__), EXCEL_FILE_PATH)
        wb = load_workbook(excel_path, data_only=True)
        ws = wb[SHEET_NAME]

        # F3 셀에 주소 입력
        ws["F3"].value = address

        # E10 계산 (현재는 단순 복사, 필요시 로직 추가)
        ws["E10"].value = ws["F3"].value
        result_value = ws["E10"].value

        result = {"E10": result_value if result_value else ""}

    except Exception as e:
        return jsonify({"error": str(e)})

    return jsonify(result)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)